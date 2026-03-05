from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from core.models import Company
from products.models import Product, Inventory


def norm(v):
    # 엑셀 값 정리(공백/특수공백 제거)
    if v is None:
        return ""
    s = str(v)
    s = s.replace("\u00a0", " ")  # NBSP
    return s.strip()


def to_int(v, default=0):
    try:
        if v is None or v == "":
            return default
        if isinstance(v, (int, float)):
            return int(v)
        s = norm(v).replace(",", "")
        if s == "":
            return default
        return int(float(s))
    except Exception:
        return default


class Command(BaseCommand):
    help = "엑셀(xlsx)로 품목(Product)과 재고(Inventory) 초기 데이터를 입력/업데이트합니다."

    def add_arguments(self, parser):
        parser.add_argument("--company-id", type=int, required=True, help="대상 회사 ID")
        parser.add_argument("--file", type=str, required=True, help="xlsx 파일 경로")
        parser.add_argument("--sheet", type=str, default=None, help="시트명(없으면 첫 시트)")
        parser.add_argument("--qty-col", type=str, default="현재고", help="수량 컬럼명(기본: 현재고)")

    @transaction.atomic
    def handle(self, *args, **options):
        company_id = options["company_id"]
        file_path = options["file"]
        sheet_name = options["sheet"]
        qty_col = options["qty_col"]

        try:
            company = Company.objects.get(id=company_id)
        except Company.DoesNotExist:
            raise CommandError("해당 company-id가 존재하지 않습니다.")

        wb = load_workbook(filename=file_path, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.worksheets[0]

        # 1행을 헤더로 가정
        headers = [norm(c.value) for c in ws[1]]
        header_map = {h: idx for idx, h in enumerate(headers) if h}

        required = ["품목코드", "품명"]
        for r in required:
            if r not in header_map and (r + " ") not in header_map:
                raise CommandError(f"엑셀 헤더에 '{r}' 컬럼이 없습니다. 현재 헤더: {headers}")

        # 헤더명에 공백이 섞여 있을 수 있어서 보정
        def hidx(name):
            return header_map.get(name, header_map.get(name + " ", None))

        sku_i = hidx("품목코드")
        name_i = hidx("품명")
        spec_i = hidx("규격")
        unit_i = hidx("단위")
        maker_i = hidx("메이커명")
        loc_i = hidx("위치")
        cost_i = hidx("재고단가")
        qty_i = hidx(qty_col)

        created_p = 0
        updated_p = 0
        touched_inv = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            sku = norm(row[sku_i]) if sku_i is not None else ""
            name = norm(row[name_i]) if name_i is not None else ""

            if not sku or not name:
                continue

            spec = norm(row[spec_i]) if spec_i is not None else ""
            unit = norm(row[unit_i]) if unit_i is not None else ""
            maker = norm(row[maker_i]) if maker_i is not None else ""
            loc = norm(row[loc_i]) if loc_i is not None else ""
            cost_price = to_int(row[cost_i], 0) if cost_i is not None else 0
            qty = to_int(row[qty_i], 0) if qty_i is not None else 0

            product, created = Product.objects.get_or_create(
                company=company,
                sku=sku,
                defaults={
                    "name": name,
                    "spec": spec,
                    "unit": unit,
                    "maker": maker,
                    "location": loc,
                    "cost_price": cost_price,
                },
            )

            if created:
                created_p += 1
            else:
                # 변경사항 업데이트
                changed = False
                for field, val in [
                    ("name", name),
                    ("spec", spec),
                    ("unit", unit),
                    ("maker", maker),
                    ("location", loc),
                ]:
                    if getattr(product, field) != val and val != "":
                        setattr(product, field, val)
                        changed = True

                if cost_price and product.cost_price != cost_price:
                    product.cost_price = cost_price
                    changed = True

                if changed:
                    product.save()
                    updated_p += 1

            inv, _ = Inventory.objects.get_or_create(company=company, product=product)
            inv.quantity = qty
            inv.save()
            touched_inv += 1

        self.stdout.write(self.style.SUCCESS(
            f"완료: Product 생성 {created_p}, Product 업데이트 {updated_p}, Inventory 반영 {touched_inv}"
        ))